import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def export_npz_to_excel(npz_file_path, excel_file_path):
    print(f"正在读取文件: {npz_file_path}")
    try:
        data = np.load(npz_file_path, allow_pickle=True)
    except FileNotFoundError:
        print(f"错误: 找不到文件 {npz_file_path}")
        return
    except Exception as e:
        print(f"加载文件时发生错误: {e}")
        return

    # 创建 Excel 工作簿
    wb = Workbook()
    wb.remove(wb.active) # 移除默认的空白 Sheet

    # 预设表格样式
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # 深蓝色表头
    header_font = Font(color="FFFFFF", bold=True)
    row_fill_light = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_fill_dark = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid") # 斑马纹
    thin_border = Border(left=Side(style='thin', color='CCCCCC'),
                         right=Side(style='thin', color='CCCCCC'),
                         top=Side(style='thin', color='CCCCCC'),
                         bottom=Side(style='thin', color='CCCCCC'))

    # 1. 创建 Summary (数据摘要) 页
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.append(["数据矩阵 (Array Name)", "形状 (Shape)", "数据类型", "最小值", "最大值", "平均值"])
    
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws_summary.column_dimensions[cell.column_letter].width = 20
    ws_summary.freeze_panes = "A2"

    for r_idx, array_name in enumerate(data.files, 2):
        arr = data[array_name]
        
        # 处理非数值型或空数组，避免统计报错
        if np.issubdtype(arr.dtype, np.number) and arr.size > 0:
            min_val, max_val, mean_val = float(np.min(arr)), float(np.max(arr)), float(np.mean(arr))
        else:
            min_val, max_val, mean_val = "N/A", "N/A", "N/A"

        row_data = [array_name, str(arr.shape), str(arr.dtype), min_val, max_val, mean_val]
        ws_summary.append(row_data)
        
        fill = row_fill_light if r_idx % 2 == 0 else row_fill_dark
        for c_idx, value in enumerate(row_data, 1):
             cell = ws_summary.cell(row=r_idx, column=c_idx)
             cell.fill = fill
             cell.border = thin_border
             cell.alignment = Alignment(horizontal='center')
             if isinstance(value, float):
                 cell.number_format = '0.0000'

    # 2. 生成具体数据页 (qpos 和 actions 等)
    for array_name in data.files:
        ws = wb.create_sheet(title=array_name)
        arr = data[array_name]
        
        # 将高维数据展平为2D以便放入表格
        if arr.ndim > 2:
            arr = arr.reshape(arr.shape[0], -1) 
            
        columns = [f"Dim_{i+1}" for i in range(arr.shape[1])] if arr.ndim == 2 else ["Value"]
        df = pd.DataFrame(arr, columns=columns)
        df.insert(0, "Time_Step", range(len(arr))) # 增加时间步索引
        
        ws.append(df.columns.tolist())
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[cell.column_letter].width = 15
        ws.freeze_panes = "A2"
        
        # 写入数据并设置斑马纹
        print(f"正在写入 {array_name} 数据至 Excel ...")
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            ws.append(row)
            fill = row_fill_light if r_idx % 2 == 0 else row_fill_dark
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                if c_idx > 1: # 对实际数据列设置小数位格式
                    cell.number_format = '0.0000'

    # 保存文件
    data.close()
    wb.save(excel_file_path)
    print(f"✅ 转换成功！文件已保存至: {excel_file_path}")

# --- 运行代码 ---
if __name__ == "__main__":
    # 使用你环境中的实际路径
    npz_file = '/home/jun/xunjian_vla_workspace/datasets/screwdriver_cleanup/ep_0/joint_data.npz'
    
    # 自动生成同目录下的 excel 文件路径
    excel_file = os.path.splitext(npz_file)[0] + '.xlsx' 
    
    export_npz_to_excel(npz_file, excel_file)